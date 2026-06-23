"""DE差分进化 + CMA-ES优化测试: 直接优化全参数空间 (拓扑+流量+H₂), 非理想。

  外部热源 Air 100kg/s→H₂冷源→He工质闭环循环, 优化器直接搜索
  t_min/t_max/t_q/p_q/H₂_mf/高斯基权重, 目标最小化 hx_unmatched(星型HX匹配)。
  每代打印最优值；输出 DE 收敛曲线 + TS/PS 图 + Excel 过程日志。
"""

from __future__ import annotations

import sys
import time
from datetime import datetime
from pathlib import Path

import pytest

from core import (
    ClosedCycleLayer,
    ClosedCycleTPInput,
    CycleConfig,
    ExternalSourceInput,
    HeatTQCurve,
    ProcessCategory,
    PropertyRegistry,
    SystemInput,
    build_heat_tq_curves,
)
import config as cyges_config
from optimize import Optimizer

TESTS_DIR = Path(__file__).resolve().parent


def _make_system_input() -> SystemInput:
    """Air/H₂/He 系统，非理想。"""
    hot = ExternalSourceInput(
        fluid="Air", mass_flow=100.0,
        T_in=1250.0, P_in=200.0, T_out=500.0, P_out=180.0,
    )
    cold = ExternalSourceInput(
        fluid="Hydrogen", mass_flow=4.3,
        T_in=20.0, P_in=5000.0, T_out=900.0, P_out=4500.0,
    )
    cycle_input = ClosedCycleTPInput(
        fluid="He", t_min=40.0, t_max=1000.0, p_min=2000.0, p_max=10000.0,
        t_quantiles=(), p_quantiles=(),
        subcycle_mass_flow_initial=20.0,
    )
    return SystemInput(
        heat_sources=(hot,), cold_sources=(cold,),
        cycles=(CycleConfig(
            input=cycle_input, use_non_ideal=True,
            delta_T_min=20.0, heat_method=None,
        ),),
        delta_T_min=20.0, heat_method="system_pinch",
    )


def _get_case_name(log_path: Path) -> str:
    """根据已有日志行数生成新工况名 run_001, run_002, ..."""
    if log_path.exists():
        import openpyxl
        wb = openpyxl.load_workbook(log_path)
        n = wb.active.max_row  # 含表头，所以行数-1 = 已有数据行
        wb.close()
        return f"run_{n:03d}"
    return "run_001"


def _append_excel_log(result, opt, sys_inp, elapsed, case_name, filepath):
    """追加一行优化日志到 Excel 文件。"""
    import openpyxl

    sp = result.system_result.system_pinch
    ct = result.system_result.cycle_reports[0].cycle_totals
    cfg = sys_inp.cycles[0]
    tp = cfg.input
    cold_src = sys_inp.cold_sources[0]

    # 源功率
    from core.fluid_property_solver import PropertyRegistry
    _p = PropertyRegistry()
    h_in = _p("Air", "TP", sys_inp.heat_sources[0].T_in, sys_inp.heat_sources[0].P_in)["H"]
    h_out = _p("Air", "TP", sys_inp.heat_sources[0].T_out, sys_inp.heat_sources[0].P_out)["H"]
    hot_power = sys_inp.heat_sources[0].mass_flow * (h_out - h_in)
    c_in = _p(cold_src.fluid, "TP", cold_src.T_in, cold_src.P_in)["H"]
    c_out = _p(cold_src.fluid, "TP", cold_src.T_out, cold_src.P_out)["H"]
    cold_power = cold_src.mass_flow * (c_out - c_in)

    # 高斯基权重
    idx = 2 + opt._n_t_q + opt._n_p_q
    basis_w = [f"{result.x_opt[idx+i]:.2f}" for i in range(opt._mf_dim)]

    # 分位（离散后）
    from optimize.solver import _round_and_dedup
    raw_t = tuple(result.x_opt[2+di] for di in range(opt._n_t_q))
    raw_p = tuple(result.x_opt[2+opt._n_t_q+pi] for pi in range(opt._n_p_q))
    tq_dedup = _round_and_dedup(raw_t, opt._qstep, opt._qmerge)
    pq_dedup = _round_and_dedup(raw_p, opt._qstep, opt._qmerge)

    # 过程计数
    n_map = {cat: 0 for cat in ProcessCategory}
    for _, rec in result.system_result.cycle_reports[0].by_edge:
        n_map[rec.category] += 1

    row = [
        case_name,
        datetime.now().isoformat(timespec="seconds"),
        round(cold_src.mass_flow, 2),
        sys_inp.heat_method,
        cfg.use_non_ideal,
        float(cyges_config.NON_IDEAL_HEAT_EFFICIENCY_DEFAULT),
        float(cyges_config.NON_IDEAL_MECHANICAL_EFFICIENCY_DEFAULT),
        f"{opt._n_basis_s}×{opt._n_basis_p}" if opt._basis else "N/A",
        opt._mf_step,
        opt._qstep,
        opt._tstep,
        opt._n_workers,
        300,  # maxiter
        42,   # seed
        f"[800,1200]/[40,200]",
        f"[{tp.p_min},{tp.p_max}]",
        opt._n_t_q,
        opt._n_p_q,
        round(result.x_opt[0], 1),
        round(result.x_opt[1], 1),
        str(tuple(round(v, 2) for v in tq_dedup)),
        str(tuple(round(v, 2) for v in pq_dedup)),
        round(result.objective, 5),
        round(sp.hot_utility_demand if sp else 0, 1),
        round(sp.cold_utility_demand if sp else 0, 1),
        round(sp.hot_utility_ratio * 100 if sp else 0, 2),
        round(sp.cold_utility_ratio * 100 if sp else 0, 2),
        round(hot_power, 1),
        round(cold_power, 1),
        round(ct.net_mechanical_power, 1),
        round(ct.net_heat_rate, 1),
        round(ct.compression_power, 1),
        round(ct.expansion_power, 1),
        round(ct.heat_absorption_rate, 1),
        round(ct.heat_rejection_rate, 1),
        n_map[ProcessCategory.COMPRESSION],
        n_map[ProcessCategory.EXPANSION],
        n_map[ProcessCategory.HEAT_ABSORPTION],
        n_map[ProcessCategory.HEAT_REJECTION],
        result.n_evaluations,
        round(elapsed, 1),
    ]

    headers = [
        "case_name",
        "timestamp", "h2_flow", "heat_method", "non_ideal", "sigma", "eta_is",
        "basis", "step_mf", "step_q", "step_t", "n_workers", "maxiter", "seed",
        "t_bounds", "p_range",
        "n_t_q", "n_p_q", "opt_t_max", "opt_t_min", "opt_t_q", "opt_p_q",
        "obj", "hot_kW", "cold_kW", "hot_pct", "cold_pct",
        "hot_src_kW", "cold_src_kW", "net_mech_kW", "net_heat_kW",
        "comp_kW", "exp_kW", "abs_kW", "rej_kW",
        "n_comp", "n_exp", "n_abs", "n_rej",
        "n_evals", "runtime_s",
    ]

    if Path(filepath).exists():
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
    ws.append(row)
    wb.save(filepath)


def test_optimize_min_max_utility() -> None:
    """DE + 高斯基编码优化。"""
    sys_inp = _make_system_input()
    props = PropertyRegistry()

    # 工况名和输出文件夹
    log_path = TESTS_DIR / "opt_log.xlsx"
    case_name = _get_case_name(log_path)
    run_dir = TESTS_DIR / case_name
    run_dir.mkdir(parents=True, exist_ok=True)
    print(f"\n工况: {case_name}\n")

    opt = Optimizer(base_input=sys_inp, props=props, objective="min_max_utility",
                    mf_step_fraction=0.01, quantile_step=0.01,
                    basis_encoding=True, basis_s=3, basis_p=3,
                    mf_bounds=(0.0, 50.0), quantile_merge_ratio=0.0,
                    n_workers=6)

    mf_label = f"{opt._mf_dim}×basis"
    print(f"\n维度: {len(opt.bounds)} (t_max + t_min + t_q + p_q + {mf_label})")
    print("CMA-ES, maxiter=300, seed=42, 高斯基编码\n")

    history: list[tuple[int, float]] = []  # (gen, best_val)

    def _on_gen(gen: int, restart: int, best_x: list[float], best_val: float, n_evals: int) -> None:
        history.append((gen, best_val))
        idx = 0
        t_max, t_min = best_x[idx], best_x[idx+1]; idx += 2
        t_q_vals = ", ".join(f"{best_x[idx+i]:.3f}" for i in range(opt._n_t_q))
        idx += opt._n_t_q
        p_q_vals = ", ".join(f"{best_x[idx+i]:.3f}" for i in range(opt._n_p_q))
        idx += opt._n_p_q
        basis_summary = f"w_mean={sum(best_x[idx:])/max(1,opt._mf_dim):.1f}"
        sys.stdout.write(
            f"  gen {gen:3d} | r{restart} | obj={best_val:.5f} | t_max={t_max:.0f} t_min={t_min:.0f}"
            f" | t_q=[{t_q_vals}] p_q=[{p_q_vals}] | {basis_summary} | evals={n_evals}\n"
        )
        sys.stdout.flush()

    print("  开始优化...\n")
    t0 = time.time()
    result = opt.run(
        method="cma", maxiter=300, seed=42, early_stop=30, sigma0=0.3,
        callback=_on_gen,
    )
    elapsed = time.time() - t0

    sp = result.system_result.system_pinch
    assert sp is not None, "系统夹点不应为空"

    report = result.system_result.cycle_reports[0]
    ct = report.cycle_totals

    print("\n" + "=" * 55)
    print("最优解")
    print("=" * 55)
    print(f"  t_max / t_min     = {result.x_opt[0]:.1f} / {result.x_opt[1]:.1f} K")
    idx = 2
    tq_vals = [f"{result.x_opt[idx+i]:.4f}" for i in range(opt._n_t_q)]
    idx += opt._n_t_q
    pq_vals = [f"{result.x_opt[idx+i]:.4f}" for i in range(opt._n_p_q)]
    idx += opt._n_p_q
    print(f"  t_q ({opt._n_t_q}个)     = [{', '.join(tq_vals)}]")
    print(f"  p_q ({opt._n_p_q}个)     = [{', '.join(pq_vals)}]")
    w_summary = ", ".join(f"{result.x_opt[idx+i]:.2f}" for i in range(min(opt._mf_dim, len(result.x_opt)-idx)))
    print(f"  basis ({opt._mf_dim}个) = [{w_summary}]")
    print(f"  objective          = {result.objective:.5f}")
    print(f"  hot_utility_demand = {sp.hot_utility_demand:.1f} kW ({sp.hot_utility_ratio*100:.2f}%)")
    print(f"  cold_utility_demand= {sp.cold_utility_demand:.1f} kW ({sp.cold_utility_ratio*100:.2f}%)")
    print(f"  net_mechanical     = {ct.net_mechanical_power:.1f} kW")
    print(f"  net_heat           = {ct.net_heat_rate:.1f} kW")
    print(f"  n_evaluations      = {result.n_evaluations}")

    assert result.n_evaluations > 0

    # ── 绘制残差下降曲线 ──
    matplotlib = pytest.importorskip("matplotlib")
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    gens, vals = zip(*history)
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(gens, vals, "o-", markersize=2, linewidth=1.0, color="tab:blue")
    ax.set_xlabel("Generation"); ax.set_ylabel("Objective (min-max utility ratio)")
    ax.set_title("DE Convergence — Air/H₂/He (non-ideal, seed=42)")
    ax.grid(True, alpha=0.25)
    ax.set_ylim(bottom=0)

    # 标注初始和最终值
    ax.annotate(f"{vals[0]:.4f}", (gens[0], vals[0]), textcoords="offset points", xytext=(5, 8), fontsize=8, color="0.3")
    ax.annotate(f"{vals[-1]:.4f}", (gens[-1], vals[-1]), textcoords="offset points", xytext=(5, -12), fontsize=8, color="0.3")

    out = run_dir / "convergence.png"
    fig.savefig(out, dpi=150, bbox_inches="tight")
    plt.close(fig)
    assert out.is_file()

    # ── 用最优参数绘图（复用 test_system_full 的 _draw_pinch_row）──
    from tests.test_system_full import _draw_pinch_row

    best_result = result.system_result
    best_report = best_result.cycle_reports[0]
    best_pinch = best_result.cycle_pinch or best_result.system_pinch
    assert best_pinch is not None

    tq_curves = list(build_heat_tq_curves(best_report, props))

    # 图 A: 循环 TS/PS（单张）
    fig_a, (ax_ts, ax_ps) = plt.subplots(1, 2, figsize=(12, 5))
    fig_a.suptitle("Optimized cycle: T–S & P–S", fontsize=11)
    for _, snap in best_report.nodes:
        ax_ts.scatter(snap.S, snap.T, s=10, color="0.3", zorder=2)
        ax_ps.scatter(snap.S, snap.P, s=10, color="0.3", zorder=2)
    for _, rec in best_report.by_edge:
        c = "tab:orange" if rec.kind == "mechanical" else "tab:green"
        ax_ts.plot([rec.tail_state.S, rec.head_state.S], [rec.tail_state.T, rec.head_state.T], color=c, lw=1.0, alpha=0.7)
        ax_ps.plot([rec.tail_state.S, rec.head_state.S], [rec.tail_state.P, rec.head_state.P], color=c, lw=1.0, alpha=0.7)
        ts_fr, ts_to = 0.38, 0.62
        for ax, y1, y2 in [(ax_ts, rec.tail_state.T, rec.head_state.T), (ax_ps, rec.tail_state.P, rec.head_state.P)]:
            ax.annotate("", xytext=(rec.tail_state.S + ts_fr*(rec.head_state.S-rec.tail_state.S), y1 + ts_fr*(y2-y1)),
                        xy=(rec.tail_state.S + ts_to*(rec.head_state.S-rec.tail_state.S), y1 + ts_to*(y2-y1)),
                        arrowprops=dict(arrowstyle="->", color="0.3", lw=1.2, alpha=0.7))
    ax_ts.set_xlabel("S"); ax_ts.set_ylabel("T [K]"); ax_ts.grid(True, alpha=0.25)
    ax_ps.set_xlabel("S"); ax_ps.set_ylabel("P [kPa]"); ax_ps.grid(True, alpha=0.25)
    fig_a.tight_layout(); fig_a.savefig(run_dir / "cycle_ts_ps.png", dpi=150); plt.close(fig_a)

    # 图 B: 循环夹点 T-Q + 组成
    curves_for_pinch: list[HeatTQCurve] = []
    if best_pinch.matched_absorption.q_points != (0.0,) or best_pinch.matched_absorption.t_points != (0.0,):
        curves_for_pinch.append(best_pinch.matched_absorption)
    if best_pinch.matched_rejection.q_points != (0.0,) or best_pinch.matched_rejection.t_points != (0.0,):
        curves_for_pinch.append(best_pinch.matched_rejection)
    if best_pinch.extra_absorption is not None:
        curves_for_pinch.append(best_pinch.extra_absorption)
    if best_pinch.extra_rejection is not None:
        curves_for_pinch.append(best_pinch.extra_rejection)
    fig_b, (ax_tq_b, ax_bar_b) = plt.subplots(1, 2, figsize=(14, 6))
    _draw_pinch_row(ax_tq_b, ax_bar_b, curves_for_pinch, "opt", pinch=best_pinch)
    fig_b.savefig(run_dir / "pinch_composition.png", dpi=150); plt.close(fig_b)

    # 图 C: 循环性能
    fig_c, (ax_mech, ax_cat) = plt.subplots(1, 2, figsize=(12, 5))
    fig_c.suptitle("Optimized cycle performance", fontsize=11)
    mech_data = []; cat_sums = {pc: 0.0 for pc in (ProcessCategory.COMPRESSION, ProcessCategory.EXPANSION, ProcessCategory.HEAT_ABSORPTION, ProcessCategory.HEAT_REJECTION)}
    for _, rec in best_report.by_edge:
        if rec.power_rate is None: continue
        cat_sums[rec.category] += rec.power_rate
        if rec.kind == "mechanical":
            signed = abs(rec.power_rate) if rec.category == ProcessCategory.COMPRESSION else -abs(rec.power_rate)
            mech_data.append((rec.edge_key, signed, abs(rec.power_rate)))
    mech_data.sort(key=lambda x: x[2])
    if mech_data:
        labels, vals_m = zip(*[(r[0], r[1]) for r in mech_data])
        colors = ["tab:blue" if v >= 0 else "tab:orange" for v in vals_m]
        ax_mech.bar(range(len(labels)), vals_m, color=colors, alpha=0.9)
        ax_mech.set_xticks(range(len(labels))); ax_mech.set_xticklabels(labels, rotation=60, ha="right", fontsize=6)
    ax_mech.axhline(0, color="0.2"); ax_mech.set_title("mechanical"); ax_mech.set_ylabel("kW"); ax_mech.grid(True, axis="y", alpha=0.25)
    cn = ["compression", "expansion", "absorption", "rejection"]
    cv = [cat_sums[pc] for pc in (ProcessCategory.COMPRESSION, ProcessCategory.EXPANSION, ProcessCategory.HEAT_ABSORPTION, ProcessCategory.HEAT_REJECTION)]
    ax_cat.bar(cn, cv, color=["tab:orange","tab:blue","tab:red","tab:green"], alpha=0.9)
    ax_cat.axhline(0, color="0.2"); ax_cat.set_title("category sums"); ax_cat.grid(True, axis="y", alpha=0.25)
    ax_cat.text(3.5, max(abs(v) for v in cv)*0.9, f"Σ mech={cv[0]+cv[1]:.0f}\nΣ heat={cv[2]+cv[3]:.0f}", fontsize=8, ha="center", bbox=dict(boxstyle="round", alpha=0.2))
    fig_c.tight_layout(); fig_c.savefig(run_dir / "performance.png", dpi=150); plt.close(fig_c)

    # 图 D: 系统夹点
    sys_pinch = best_result.system_pinch
    if sys_pinch is not None:
        sys_curves: list[HeatTQCurve] = [sys_pinch.matched_absorption, sys_pinch.matched_rejection]
        if sys_pinch.extra_absorption is not None:
            sys_curves.append(sys_pinch.extra_absorption)
        if sys_pinch.extra_rejection is not None:
            sys_curves.append(sys_pinch.extra_rejection)
        sys_curves = [c for c in sys_curves if not (c.q_points == (0.0,) and c.t_points == (0.0,))]
        fig_d, (ax_tq_d, ax_bar_d) = plt.subplots(1, 2, figsize=(14, 6))
        _draw_pinch_row(ax_tq_d, ax_bar_d, sys_curves, "opt system", pinch=sys_pinch)
        fig_d.savefig(run_dir / "system_pinch.png", dpi=150); plt.close(fig_d)

    # 图 E: 子循环网格 + 流量（TS/PS, 原拓扑, 参数与优化器对齐）
    idx = 0
    t_max_opt, t_min_opt = result.x_opt[0], result.x_opt[1]; idx = 2
    raw_t = tuple(result.x_opt[idx+i] for i in range(opt._n_t_q)); idx += opt._n_t_q
    raw_p = tuple(result.x_opt[idx+i] for i in range(opt._n_p_q))
    # 应用相同的离散化
    from optimize.solver import _round_and_dedup
    t_q_opt = _round_and_dedup(raw_t, opt._qstep, opt._qmerge)
    p_q_opt = _round_and_dedup(raw_p, opt._qstep, opt._qmerge)

    tp_opt = ClosedCycleTPInput(
        fluid="He", t_min=t_min_opt, t_max=t_max_opt,
        p_min=opt._base_tp.p_min, p_max=opt._base_tp.p_max,
        t_quantiles=t_q_opt, p_quantiles=p_q_opt,
    )
    layer_opt = ClosedCycleLayer(tp_opt)
    weights = list(result.x_opt[4:4+opt._mf_dim])
    if opt._basis:
        mf_opt = opt._decode_basis_flows(layer_opt, weights)
    else:
        mf_opt = [float(result.x_opt[4+i]) for i in range(min(len(layer_opt.subcycles), len(result.x_opt)-4))]
    layer_opt.subcycle_mass_flows = mf_opt
    layer_opt.commit_subcycle_mass_flows_to_topology()

    fig_e, (ax_ts_e, ax_ps_e) = plt.subplots(1, 2, figsize=(14, 6))
    fig_e.suptitle("Optimized subcycle grid & flows", fontsize=11)

    # 散点: 所有节点
    for n in layer_opt.nodes.values():
        ax_ts_e.scatter(n.S, n.T, s=8, color="0.4", zorder=1)
        ax_ps_e.scatter(n.S, n.P, s=8, color="0.4", zorder=1)

    # 子循环多边形 + 流量标注
    colors_sc = plt.cm.tab10.colors
    for i, sc in enumerate(layer_opt.subcycles):
        c = colors_sc[i % len(colors_sc)]
        q = sc.mass_flow if sc.mass_flow is not None else 0.0
        n0, n1, n2, n3 = [layer_opt.nodes[idx] for idx in sc.nodes]
        for ax, y_get in [(ax_ts_e, lambda n: n.T), (ax_ps_e, lambda n: n.P)]:
            xs = [n0.S, n1.S, n2.S, n3.S, n0.S]
            ys = [y_get(n0), y_get(n1), y_get(n2), y_get(n3), y_get(n0)]
            ax.plot(xs, ys, "-", color=c, linewidth=1.2, alpha=0.5)
            cx = (n0.S + n2.S) / 2
            cy = (y_get(n0) + y_get(n2)) / 2
            ax.annotate(f"SC{i}\n{q:.1f}", (cx, cy), ha="center", va="center", fontsize=6, color=c, fontweight="bold")

    ax_ts_e.set_xlabel("S [kJ/(kg·K)]"); ax_ts_e.set_ylabel("T [K]"); ax_ts_e.grid(True, alpha=0.25)
    ax_ps_e.set_xlabel("S [kJ/(kg·K)]"); ax_ps_e.set_ylabel("P [kPa]"); ax_ps_e.grid(True, alpha=0.25)
    fig_e.tight_layout(); fig_e.savefig(run_dir / "subcycle_grid.png", dpi=150); plt.close(fig_e)

    print(f"  图像已保存至: {run_dir}/")

    # ── Excel 日志记录 ──
    _append_excel_log(result, opt, sys_inp, elapsed=elapsed, case_name=case_name, filepath=log_path)
    print(f"  Excel 日志已追加: {log_path.name}")